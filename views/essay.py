from django.shortcuts import render, redirect
from app_01 import models

from app_01.utils.pagination import Pagination
from app_01.utils.form import EssayModelForm, EssayEditModelForm


def essay_list(request):

    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["Ref_DOI_number__contains"] = search_data

    queryset = models.Essay.objects.filter(**data_dict).order_by("-Ref_Publication_date")

    page_object = Pagination(request, queryset)

    context = {
        "search_data": search_data,

        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 页码
    }
    return render(request, 'essay_list.html', context)


def essay_add(request):

    if request.method == "GET":
        form = EssayModelForm()
        return render(request, 'essay_add.html', {"form": form})
    form = EssayModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/essay/list/')
    return render(request, 'essay_add.html', {"form": form})



def essay_delete(request):
    """ 删除部门 """
    # 获取ID http://127.0.0.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')

    # 删除
    models.Essay.objects.filter(id=nid).delete()

    # 重定向回部门列表
    return redirect("/essay/list/")


def essay_edit(request, nid):

    row_object = models.Essay.objects.filter(id=nid).first()

    if request.method == "GET":
        form = EssayEditModelForm(instance=row_object)
        return render(request, 'essay_edit.html', {"form": form})

    form = EssayEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/essay/list/')

    return render(request, 'essay_edit.html', {"form": form})


def essay_multi(request):
    """ 批量删除（Excel文件）"""
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        # text = row[0].value
        # print(text)
        list0 = row[0].value    #Ref_Publication_date
        list1 = row[1].value    #Ref_Publication_date_digit
        list2 = row[2].value    #Ref_Article_title
        list3 = row[3].value    #Ref_Journal
        list4 = row[4].value    #Ref_First_author
        list5 = row[5].value    #Ref_Corresponding_author
        list6 = row[6].value    #Ref_DOI_number
        list7 = row[7].value    #TENG_MFCFP
        list8 = row[8].value    #TENG_structure
        list9 = row[9].value    #Structure
        list10= row[10].value   #TENG_Mode
        list11 = row[11].value  #Mode
        list12 = row[12].value  #Mode_Loss_electrons_Electrode_layer
        list13 = row[13].value  #Mode_Loss_electrons_layer
        list14 = row[14].value  #Mode_Get_electrons_layer
        list15 = row[15].value  #Mode_Get_electrons_layer_Electrode_layer
        list16 = row[16].value  #TENG_Power_generation_application
        list17 = row[17].value  #TENG_Sensor_field_application
        list18 = row[18].value  #Efficiency_Voltage
        list19 = row[19].value  #Efficiency_Current
        list20 = row[20].value  #Efficiency_Current_nA
        list21 = row[21].value  #Efficiency_Current_modified
        list22 = row[22].value  #Efficiency_Charge
        list23 = row[23].value  #Efficiency_Charge_nC
        list24 = row[24].value  #Efficiency_Charge_modified
        list25 = row[25].value  #Efficiency_Power
        list26 = row[26].value  #Efficiency_Power_modified
        list27 = row[27].value  #Efficiency_Resistance
        list28 = row[28].value  #Efficiency_Area
        list29 = row[29].value  #Level_choices
        #print(list0)
        #print(list1)
        #print(list2)
        #print(list6)
        #print(list29)
        #list3 = [list(a) for a in zip(list1, list2)]
        #print(list3)
        exists = models.Essay.objects.filter(Ref_DOI_number=list6).exists()
        if not exists:
        #     print(text)
            models.Essay.objects.create(Ref_Publication_date=list0, Ref_Publication_date_digit=list1, Ref_Article_title=list2,
                                    Ref_Journal=list3, Ref_First_author=list4,Ref_Corresponding_author=list5,
                                    Ref_DOI_number=list6, TENG_MFCFP=list7,TENG_structure=list8,
                                    Structure=list9, TENG_Mode=list10, Mode=list11,Mode_Loss_electrons_Electrode_layer=list12,
                                    Mode_Loss_electrons_layer=list13, Mode_Get_electrons_layer=list14, Mode_Get_electrons_layer_Electrode_layer=list15,
                                    TENG_Power_generation_application=list16,TENG_Sensor_field_application=list17, Efficiency_Voltage=list18, Efficiency_Current=list19,
                                    Efficiency_Current_nA=list20, Efficiency_Current_modified=list21, Efficiency_Charge=list22,Efficiency_Charge_nC=list23,
                                    Efficiency_Charge_modified=list24, Efficiency_Power=list25,Efficiency_Power_modified=list26, Efficiency_Resistance=list27,
                                    Efficiency_Area=list28, Level=list29,
                                        )

    return redirect('/essay/list/')